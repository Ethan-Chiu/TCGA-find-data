#---------------------------------------------------------------------------------------
import os.path
import openpyxl
from os import listdir
from os.path import isfile, isdir, join
from pathlib import Path

#---------------------------------------------------------------------------------------
#input
print("Drag in the file: result excel file:")
result_file_name = os.path.basename(input()).rstrip()
wb = openpyxl.load_workbook(result_file_name)
sheet = wb.active

print("Drag in the file: compare result excel file:")
Cresult_file_name = os.path.basename(input()).rstrip()
Cwb = openpyxl.load_workbook(Cresult_file_name)
Csheet = Cwb.active

sheetrowmax = sheet.max_row - 1
Csheetrowmax = Csheet.max_row - 1

for i in range(1, Csheetrowmax + 1):
  for j in range(1, sheetrowmax + 1):
    if str(sheet.cell(column = 1, row = j).value).find(str(Csheet.cell(column = 1, row = i).value))>=0:
      sheet.delete_rows(j)
      break

wb.save(result_file_name)        