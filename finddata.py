#---------------------------------------------------------------------------------------
import os.path
import openpyxl
from os import listdir
from os.path import isfile, isdir, join
from pathlib import Path

#---------------------------------------------------------------------------------------
#input
print("Drag in the file: result excel file:")
result_file_name = os.path.basename(input()).rstrip()
print("Drag in the file: the matadata excel file:")
matadata_file_name = os.path.basename(input()).rstrip()
print("Drag in the file: the excel file that tells me what miR I need to find:")
what_tf_file_name = os.path.basename(input()).rstrip()
print("Drag in the file: clinical excel file:")
clinical_file_name = os.path.basename(input()).rstrip()
print("Drag in the directory: downloaded dir")
path = os.path.basename(input()).rstrip()
#---------------------------------------------------------------------------------------
#all workbook and worksheet
#4 parts need to be changed (result_file_name = 最終資料, wbmata = matadata, wtfwb = 需要找的miR, cliwb = clinical)

wb = openpyxl.load_workbook(result_file_name)
wbmeta = openpyxl.load_workbook(matadata_file_name)
wtfwb = openpyxl.load_workbook(what_tf_file_name)
cliwb = openpyxl.load_workbook(clinical_file_name)

sheet = wb.active
metasheet = wbmeta.active
wtfws = wtfwb.active
cliws = cliwb.active

#---------------------------------------------------------------------------------------
#variable
#for the position relation between case_id and filename in metadata
Fname_t_Cid = 0
#---------------------------------------------------------------------------------------
#Fname_t_Cid set
lFname = 1
lCid = 1
lFnamefind = 0
lCidfind = 0
while lFnamefind == 0:
  if str(metasheet.cell(column = 1, row = lFname).value).find(".txt")>=0:
    lFnamefind = 1
    break
  lFname +=1
  #print(lFname)
while lCidfind == 0:
  if str(metasheet.cell(column = 1, row = lCid).value).find("case_id")>=0:
    lCidfind = 1
    break
  lCid +=1
  #print(lCid)
Fname_t_Cid = lCid - lFname
#---------------------------------------------------------------------------------------
#print(Fname_t_Cid)
# 指定要列出所有檔案的目錄

# 取得所有檔案與子目錄名稱
files = listdir(path)
row = 1
for f in files:
  # 產生檔案的完整路徑
  fullpath = join(path, f)
  if isdir(fullpath):
    filefullpath = listdir(fullpath)
    for filename in filefullpath:

      if filename.find("annotations")<0:

        sheet.cell(column = 1, row = row, value = filename)

        for i in range(1, metasheet.max_row+1):
          #print(str(metasheet.cell(column = 1, row = i).value))
          if str(metasheet.cell(column = 1, row = i).value).find(filename)>=0:
            sheet.cell(column = 2, row = row, value = metasheet.cell(column = 1, row = i+Fname_t_Cid).value)
        row = row+1

wb.save(result_file_name)
#---------------------------------------------------------------------------------------
#constant numbers and variables'''
sheetrowmax = sheet.max_row

#---------------------------------------------------------------------------------------
#miRNA表現量
#funtions
def get_file_path(name):
  if (len(name) < 3):
    return ""
  for filename in Path('.').rglob(name):
    return filename

def grep_pat (filename, pattern, crow, ccol):
  if (filename == ""):
    return
  if pattern  =="":
    return
  with open(filename) as file:
    content = file.readlines()
    fnd = False
#---------------------------------------------------------------------------------------
  for row in content:
    if (pattern in row):
      rowsplt = row.split("	")
      if (pattern == rowsplt[0]):
        sheet.cell(column = ccol, row = crow, value = row.split()[1])
        fnd = True 
  if (fnd == False):
    sheet.cell(column = ccol, row = crow, value = "null")


indexrow = 1
indexcol = 10

for i in range(1, wtfws.max_row +1):
  #print(wtfws.max_row)
  pattern = str(wtfws.cell(column = 1, row = i).value)
  print(pattern)
#print(sheet.max_row)
  for j in range(1, sheetrowmax+1):
    #print(str(sheet.cell(column = 1, row = j).value))
    name = get_file_path(str(sheet.cell(column = 1, row = j).value))
    grep_pat(name, pattern, indexrow, indexcol)
    indexrow = indexrow + 1
  sheet.cell(column = indexcol, row = indexrow, value = pattern)
  indexrow = 1
  indexcol = indexcol + 1

wb.save(result_file_name)
#---------------------------------------------------------------------------------------
#clinical
#variables
cid_t_tumorS = 0
cid_t_vitalS = 0
cid_t_dtlf = 0

#lcid = 1 ---0
#ltumorS = 1 ---1
#lvitalS = 1 ---2
#ldtlf = 1 ---3
l = [1,1,1,1]
#cidfind = 0
#tumorSfind = 0
#vitalSfind = 0
#dtlffind = 0
datafind = [0,0,0,0]
strfind = ["case_id","tumor_s","vital_s","last_fol"]
#---------------------------------------------------------------------------------------
#funtions
def upsearch(i, findstr):
  finded = 0
  i +=1
  while finded == 0:
    if str(cliws.cell(column = 1, row = i).value).find(findstr)>=0:
      finded = 1
      break 
    i +=1
  return i

def downsearch(i, findstr):
  finded = 0
  i -=1
  while finded == 0:
    if str(cliws.cell(column = 1, row = i).value).find(findstr)>=0:
      finded = 1
      break 
    i -=1
  return i
#---------------------------------------------------------------------------------------
#判斷相對位置
for i in range(0,4):
  while datafind[i] == 0:
    if str(cliws.cell(column = 1, row = l[i]).value).find(strfind[i])>=0:
      datafind[i] = 1
      break
    l[i] += 1

cid_t_tumorS = l[1] - l[0]
cid_t_vitalS = l[2] - l[0]
cid_t_dtlf = l[3] -l[0]
#---------------------------------------------------------------------------------------
#search and fill in
for j in range(1, sheetrowmax+1):
  sheetcli = str(sheet.cell(column = 2, row = j).value).split('"')
  #print(sheetcli[3])
  for i in range(1, cliws.max_row+1):
    if str(cliws.cell(column = 1, row = i).value).find(sheetcli[3])>=0:
      #找到caseid在第i行
      if cid_t_tumorS > 0:
        tumorstage = str(cliws.cell(column = 1, row = upsearch(i, strfind[1])).value).split('"')
      else:
        tumorstage = str(cliws.cell(column = 1, row = downsearch(i, strfind[1])).value).split('"')
      sheet.cell(column = 4, row = j, value = tumorstage[3])
      if cid_t_vitalS > 0:
        alivedead = str(cliws.cell(column = 1, row = upsearch(i, strfind[2])).value).split('"')
      else:
        alivedead = str(cliws.cell(column = 1, row = downsearch(i, strfind[2])).value).split('"')
      sheet.cell(column = 5, row = j, value = alivedead[3])
      if alivedead[3] == "Alive":
        sheet.cell(column = 7, row = j, value = "1")
      else:
        sheet.cell(column = 7, row = j, value = "0")
      if cid_t_dtlf > 0:
        daystlstfollow = str(cliws.cell(column = 1, row = upsearch(i, strfind[3])).value)
      else:
        daystlstfollow = str(cliws.cell(column = 1, row = downsearch(i, strfind[3])).value)
      numofdays = ''.join([x for x in daystlstfollow if x.isdigit()])
      if numofdays == "":
        sheet.cell(column = 6, row = j, value = "null")
      else:
        sheet.cell(column = 6, row = j, value = numofdays)
sheet.cell(column = 4, row = sheetrowmax+1, value = "tumor stage")
sheet.cell(column = 5, row = sheetrowmax+1, value = "tumor stage")
sheet.cell(column = 6, row = sheetrowmax+1, value = "days to last follow up")
sheet.cell(column = 7, row = sheetrowmax+1, value = "0/1")
wb.save(result_file_name)